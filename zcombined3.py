from email import message
from tkinter import *
import datetime 
import time
from typing import Counter
import keyboard as k
import pyautogui
import pywhatkit as py
import smtplib
import schedule
import pandas as pd
from openpyxl import load_workbook
from tkinter import messagebox
import sys
from bs4 import BeautifulSoup
import requests
from datetime import timezone
import pytz    
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
import re

#*****************************************************************************************************************#
#New registration page GUI

def NewEntry(*args):
    email=EmailValue.get()
    phone=PhoneValue.get()
    dataset=pd.read_excel('Recipients.xlsx')
    #this helps us put email column into an numpy array
    emaildataset=dataset.iloc[:,-2].values
    phonedataset=dataset.iloc[:-1]
    row=len(emaildataset)+2
    regex = r'\b[A-Za-z0-9._%+-]+@[A-Za-z0-9.-]+.[A-Z|a-z]{2,}\b'

    if (re.fullmatch(regex, email) and email not in emaildataset):
        wb=load_workbook("Recipients.xlsx")
        ws=wb.worksheets[0]#this is the sheet number
        ws_tables=[]
        ws[f"B{row}"]=email    
        if (phone not in phonedataset):
            ws[f"C{row}"]=phone
        wb.save('Recipients.xlsx')
        newWindow.destroy()
    else:
        messagebox.showinfo("Invalid Email","The email you've entered is invalid. Please enter another email.")
    newWindow.attributes('-topmost', 1)
    newWindow.attributes('-topmost', 0)
        
    
    
def LimitPhoneDigits(*args):
    value=PhoneValue.get()
    if len(value) > 2: PhoneValue.set(value[:10])
    for i in range (0,len(value)):
        try:
            int(value[i])
        except:
            PhoneValue.set("") 

def LimitEmailDigits(*args):
    value=EmailValue.get()             

def openNewWindow():
    # Toplevel object which will
    # be treated as a new window
    global newWindow
    newWindow = Toplevel(root)
    # sets the title of the Toplevel widget
    newWindow.title("New recepient registration")
    # sets the geometry of toplevel
    newWindow.geometry("756x382")
    #next 3 commands lock the window size for aesthetic reasons
    newWindow.resizable(0, 0)
    newWindow.maxsize(756,382)
    newWindow.minsize(756,382)
    # setting background by canvas method
    c=Canvas(newWindow,bg="gray16",height=200,width=200)
    background_label=Label(newWindow,image=bgp)
    background_label.place(x=0,y=0,relwidth=1,relheight=1)
    registerbtn=Button(newWindow,command=NewEntry,text="register",borderwidth=0,activebackground="#1651A7",fg="white",bg="#1651A7",activeforeground="black",font=("Source Sans Pro SemiBold",16,"bold"))
    registerbtn.place(x=540,y=290)
    global EmailValue
    global PhoneValue
    EmailValue=StringVar()
    EmailValue.trace('w', LimitEmailDigits)
    PhoneValue=StringVar()
    PhoneValue.trace('w',LimitPhoneDigits)
    emailtext=Entry(newWindow,width=44,borderwidth=0,font=("Source Sans Pro SemiBold",14),textvariable=EmailValue)
    emailtext.place(x=85,y=203)
    phonetext=Entry(newWindow,width=27,borderwidth=0,font=("Source Sans Pro SemiBold",14),textvariable=PhoneValue)
    phonetext.place(x=85,y=300)
    newWindow.bind('<Return>',NewEntry)#pressing enter registers details

#*****************************************************************************************************************#
#Main GUI's code
def run(*args):
    global ghour
    global gmin
    ghour=HourValue.get()
    print(ghour,"this is the value of hour while exit")
    gmin=MinValue.get()
    print(gmin,"this is the value of min while exit")
    root.destroy()
   
	#this function performs action on clicking the runbtn
	
def LimitHourDigits(*args):
	#this function limits num of hour chars to 2 and sets default value to 00
    value = HourValue.get()
    if len(value) > 2: HourValue.set(value[:2])
    try:
        if value!="":
            if int(value)>23: HourValue.set("00")
    except:
        HourValue.set("00") 
    
def LimitMinDigits(*args):
	#this function limits num of min chars to 2 and sets default values to 00
    value = MinValue.get()
    if len(value) > 2: MinValue.set(value[:2])
    try:
        if value!="":
            if (int(value)>59): MinValue.set("00")	
    except:
        MinValue.set("00")
    


#this is creating a new window with a background image
root=Tk()
root.title('eNotify')#setting up the window title
root.geometry("950x601")#window size
root.resizable(0, 0)
root.maxsize(950,601)
root.minsize(950,601)#to maintain aspect ratio for aesthetic reasons
c=Canvas(root,bg="gray16",height=200,width=200)

#picking a background for root window
filename=PhotoImage(file="Data\eNotifyGUI.png")
background_label=Label(root,image=filename)
background_label.place(x=0,y=0,relwidth=1,relheight=1)

#creating run button
runbtn=Button(root,text="   run   ",command=run,bg="#1651A7",activebackground="#1651A7",fg="white",borderwidth=0,font=("Source Sans Pro SemiBold",13,"bold"))
runbtn.place(x=720,y=531)

#new student button
contactbtn=Button(root,text="  Click here  ",command=openNewWindow,activeforeground="orange",activebackground="white",bg="white",borderwidth=0,font=("Source Sans Pro SemiBold",12))
contactbtn.place(x=696,y=337)

bgp=PhotoImage(file="Data\Regpage.png")#globalvar for new window bg

HourValue = StringVar()
HourValue.trace('w', LimitHourDigits)
MinValue = StringVar()
MinValue.trace('w', LimitMinDigits)
#this is textbox for hour
hourtext=Entry(root,width=2,borderwidth=0,font=("Source Sans Pro SemiBold",15,"bold"),textvariable=HourValue)
hourtext.place(x=690,y=260)
#this is textbox for minutes
mintext=Entry(root,width=2,borderwidth=0,font=("Source Sans Pro SemiBold",15,"bold"),textvariable=MinValue)
mintext.place(x=809,y=260)

def on_closing():
    if messagebox.askokcancel("Quit", "Do you want to quit?"):
        sys.exit()

root.protocol("WM_DELETE_WINDOW", on_closing)

c.pack()
root.mainloop()

#******************************************************************************************************************#
#NOTIFICATION SENDING FUNCTIONS START HERE

def email_reminder():
  message=extract_advanced()
  email_notif(emaildataset,message[2])

def email_notif(emaildataset,message1):
  if message1:  
    with smtplib.SMTP('smtp.gmail.com',587) as smtp:
        smtp.ehlo() 
        smtp.starttls()
        smtp.ehlo()
        pwd1='yourpw'
        user1='yourmail'
        smtp.login(user=user1,password=pwd1)
        #Contents of email and concatenating them
        subject="Today's reminder"
        message=MIMEMultipart("alternative")
        message.attach(MIMEText(message1, 'html'))
        '''msg=f'Subject: {subject}\n\n{message}'''
        #sending the mail
        for email in emaildataset:
            message['From'] = 'Apsitenotify@gmail.com' 
            message['To'] = email
            message['Subject'] = "Today's Reminder"
            time.sleep(5)
            smtp.sendmail('ApsiteNotify@gmail.com',email,message.as_string())
            print("email sent to",email)
            

    
def whatsapp_notif(num,message,hour,minutes):
    py.sendwhatmsg(f"+91{num}",message,hour,minutes,wait_time=35,tab_close=False)
    pyautogui.click(1050, 950)
    time.sleep(2)
    k.press_and_release('enter')
    py.close_tab(2)

def whatsapp_numloop(num,message,currenthour):
    for i in num:
        try:
            i=int(i)
            now=datetime.datetime.now()
            hour=int(now.strftime("%H"))
            minutes=int(now.strftime("%M"))
            print(hour,minutes)
            if hour>currenthour+3:
                print('exiting main loop')
                break
            else:
                if minutes>58:
                    whatsapp_notif(i,message,hour+1,2)
                else:
                    try:
                        whatsapp_notif(i,message,hour,minutes+1)
                    except:
                        whatsapp_notif(i,message,hour,minutes+3)    
                print(f"update sent to :{i} at {hour} hours and {minutes} minutes")
        except:
            continue
#******************************************************************************************************************#
#data extraction
#RECENT ACTIVITY CODE

'''def getRecent():
  payload = {
    'username': 'user',
    'password': 'pw'
    }
  recent_string = ''
  with requests.Session() as p:
    login_response2 = p.post('http://moodle.apsit.org.in/moodle/login/index.php', data = payload)
    response3 = p.get('http://moodle.apsit.org.in/moodle/course/index.php?categoryid=409') #only for div B
    bsObj4 = BeautifulSoup(response3.text, 'lxml')
    all_course_url = bsObj4.find_all("h3", class_ = "coursename")
    for i in range(len(all_course_url)):
      course_url = all_course_url[i].a.get('href')
      course_name_recent = all_course_url[i].a.text
      response4 = p.get(course_url)
      bsObj5 = BeautifulSoup(response4.text, 'lxml') 
      try:
        recent_activity = bsObj5.find("p", class_ = "activity")
        recent_upload = recent_activity.a.text
        recent_upload_url = recent_activity.a.get('href')
        recent_string += (f'{recent_upload} is uploaded in the course {course_name_recent} and can be accessed through {recent_upload_url}\n')
      except:
        continue
  return (recent_string)'''


def extract_advanced():
  payload = {
    'username': 'user',
    'password': 'pw@Apsit'
  }
  two_day_list = ""
  alert_string = ''
  whatsapp_string = "Upcoming Reminders:\n"
  tz_IND = pytz.timezone('Asia/Kolkata')   
  datetime_IND = datetime.datetime.now(tz_IND)  
  date_now = int(datetime_IND.strftime("%d"))
  with requests.Session() as r:
    login_res = r.post('http://moodle.apsit.org.in/moodle/login/index.php', data = payload)
    upcoming_url = 'http://moodle.apsit.org.in/moodle/calendar/view.php'
    submissions = r.get(upcoming_url)
    bsObj = BeautifulSoup(submissions.text, "lxml")
    divs_of_sub = bsObj.find_all("div", class_ = "event")
    for i in range(len(divs_of_sub)):
      submission_course = divs_of_sub[i].div.a.text[6: -12]
      submission_name = divs_of_sub[i].h3.a.text
      submission_link = divs_of_sub[i].h3.a.get('href')
      submission_time_date = divs_of_sub[i].span.text
      submission_date_extract = re.findall('\d+', submission_time_date)
      submission_date = int(submission_date_extract[0])
      alert_string += '<html><head></head><body>'
      alert_string += f'<p><h2>Course name: {submission_course}</h2> has a submission named <b>{submission_name}</b> which is due on <b>{submission_time_date}</b> and can be accessed through <a href = "{submission_link}">this link.</a> </p><br><hr><br>\n'
      whatsapp_string += f"Course name: {submission_course} has a submission named {submission_name} which is due on {submission_time_date} and is accessible via this link: {submission_link}\n"
      if submission_date <= date_now + 2 and (submission_date >= date_now and submission_date <= 30):
        two_day_list=alert_string
  alert_string+='<i style="color:grey;text-align: center;font-family: sans-serif;font-size:0.7rem ">Created by <a href="https://buildforacause.github.io/team/">students</a> of APSIT</i></body></html>\n'
  return (alert_string, whatsapp_string,two_day_list)



#******************************************************************************************************************#
#MAIN CODE

def main():
    try:
        #we are extracting students info using panda
        dataset=pd.read_excel('Data\Recipients.xlsx')
        #this helps us put email column into an numpy array
        global emaildataset 
        emaildataset=dataset.iloc[:,-2].values
        whatsappdataset=dataset.iloc[:,-1].values   
        currenttime=datetime.datetime.now()
        currenthour=int(currenttime.strftime("%H"))
        #enter the number data here
        emailmessage,whatsappmessage,two_day_list=extract_advanced()
        #link data extracted message here
        email_notif(emaildataset,emailmessage)#first sending all the emails
        whatsapp_numloop(whatsappdataset,whatsappmessage,currenthour)
        #whatsapp_numloop(num,message,currenthour)  #sending all the whatsapp texts next
        print('messages sent')
        schedule.every(1).minutes.do(email_reminder)
    except:
        print("Message wasnt delivered")

#managing ghour and gmin to match the format
if len(ghour)==1:
    ghour='0'+ghour
if len(gmin)==1:
    gmin='0'+gmin    
if ghour=="":
    ghour="00"
if gmin=="":
    gmin="00"    
#*******************************************************************************************************************#

timer=ghour+":"+gmin+":00"
print("starting script at",timer)
schedule.every().day.at(timer).do(main)
#this is the main loop for every 24 hour checking
while True:
    schedule.run_pending()
    time.sleep
